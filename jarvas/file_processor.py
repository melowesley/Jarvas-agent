"""Processa arquivos (PDF/Excel/Word/CSV/imagem) guiado pelo prompt do usuario."""
from __future__ import annotations
import csv as _csv
from pathlib import Path

from jarvas.supabase_client import save_attachment

_OUTPUT_BASE = "jarvas_outputs"

_TYPE_FOLDERS: dict[str, str] = {
    ".xlsx": "excel", ".xls": "excel",
    ".csv": "csv",
    ".pdf": "pdf",
    ".docx": "docs", ".txt": "docs",
    ".jpg": "images", ".jpeg": "images", ".png": "images",
}


def _get_output_dir(ext: str, project_base: str | None) -> Path:
    folder = _TYPE_FOLDERS.get(ext.lower(), "misc")
    base = Path(project_base) if project_base else Path.home()
    out = base / _OUTPUT_BASE / folder
    out.mkdir(parents=True, exist_ok=True)
    return out


def extract_content(path: str) -> str:
    """Extrai conteudo textual do arquivo conforme tipo."""
    p = Path(path)
    if not p.exists():
        return f"[erro] arquivo nao encontrado: {path}"
    ext = p.suffix.lower()

    if ext == ".pdf":
        import fitz
        doc = fitz.open(str(p))
        return "\n".join(page.get_text() for page in doc)

    if ext in (".xlsx", ".xls"):
        import openpyxl
        wb = openpyxl.load_workbook(str(p), read_only=True, data_only=True)
        ws = wb.active
        rows = [
            "\t".join("" if c is None else str(c) for c in row)
            for row in ws.iter_rows(values_only=True)
        ]
        return "\n".join(rows)

    if ext == ".csv":
        with open(str(p), newline="", encoding="utf-8-sig") as f:
            reader = _csv.reader(f)
            return "\n".join("\t".join(row) for row in reader)

    if ext == ".docx":
        from docx import Document
        doc = Document(str(p))
        return "\n".join(para.text for para in doc.paragraphs)

    if ext == ".txt":
        return p.read_text(encoding="utf-8")

    if ext in (".jpg", ".jpeg", ".png"):
        import pytesseract
        from PIL import Image
        img = Image.open(str(p))
        return pytesseract.image_to_string(img, lang="por+eng")

    return f"[erro] tipo nao suportado: {ext}"


def _write_output(data: str, source_path: str, project_base: str | None,
                  output_format: str) -> str:
    source = Path(source_path)
    out_dir = _get_output_dir(f".{output_format}", project_base)
    out_path = out_dir / f"{source.stem}_resultado.{output_format}"

    if output_format == "xlsx":
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        for i, line in enumerate(data.splitlines(), start=1):
            cols = line.split("\t") if "\t" in line else [line]
            for j, val in enumerate(cols, start=1):
                ws.cell(row=i, column=j, value=val)
        wb.save(str(out_path))
    else:
        out_path.write_text(data, encoding="utf-8")

    return str(out_path)


def process_file(
    path: str,
    instruction: str,
    project_base: str | None,
    session_id: str,
) -> dict:
    """Extrai conteudo do arquivo, processa com IA e salva resultado."""
    import os
    from jarvas.hermes_client import chat as hermes_chat
    from jarvas.file_editor import find_file_in_project

    # Resolucao case-insensitive: se o path nao existe diretamente, busca no projeto
    resolved = path
    if not Path(path).exists() and project_base:
        found = find_file_in_project(Path(path).name, project_base)
        if found:
            resolved = found
        else:
            available = ", ".join(os.listdir(project_base)) if os.path.isdir(project_base) else "—"
            return {
                "error": (
                    f"'{Path(path).name}' não encontrado em {project_base}\n"
                    f"Arquivos disponíveis: {available}"
                )
            }

    content = extract_content(resolved)
    if content.startswith("[erro]"):
        return {"error": content}

    prompt = (
        f"Arquivo: {Path(resolved).name}\n"
        f"Conteudo extraido:\n{content[:8000]}\n\n"
        f"Instrucao: {instruction}\n\n"
        "Execute a instrucao. Para dados tabulares, use formato TSV "
        "(colunas separadas por tab) para facilitar exportacao."
    )
    resp, _ = hermes_chat(prompt)

    instr_lower = instruction.lower()
    if "excel" in instr_lower or "xlsx" in instr_lower:
        out_format = "xlsx"
    elif "csv" in instr_lower:
        out_format = "csv"
    else:
        out_format = "txt"

    out_path = _write_output(resp, resolved, project_base, out_format)

    save_attachment(
        session_id,
        Path(resolved).name,
        Path(resolved).suffix.lstrip("."),
        content[:2000],
        resp[:2000],
    )

    return {
        "output_path": out_path,
        "summary": resp[:500],
        "file_type": Path(resolved).suffix,
    }
